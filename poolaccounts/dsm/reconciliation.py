
import json
from datetime import datetime
from django.http import JsonResponse,HttpResponse , FileResponse
from dsm.common import add530hrstoDateString, format_indian_currency,get_month_start_end_dates
from registration.extarctdb_errors import extractdb_errormsg
from openpyxl import load_workbook
from registration.models import Registration
from django.db.models import Q
from .models import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os , random
import pandas as pd
from poolaccounts.settings import base_dir
from dsm.user_recon import checkBillsNotified
from .reconciliation import *
from .user_recon import *

def getPrevYearMonth(selected_month):
    # get closing balances of prev month (selected_monht in '2024-10' format)
    prev_monthyear_splt = selected_month.split('-')
    year , month = int(prev_monthyear_splt[0]) , int(prev_monthyear_splt[1])
    if month > 1:
        actual_month = month-1
        actual_month = '0'+ str(actual_month) if actual_month < 10 else actual_month
        actual_year = year
    else:
        actual_month = 12
        actual_year = year - 1

    prev_month_year = str(actual_year)+'-'+str(actual_month)

    return prev_month_year

def prepareSummarySheet(wb , summary_sheet):
    try:
        # finally add the Summary sheet
        sheet_title= 'Summary'
        ws_new = wb.create_sheet(title=sheet_title)
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Populate the sheet and apply styles
        for row_index, row_data in enumerate(summary_sheet, start=1):
            for col_index, value in enumerate(row_data, start=1):
                cell = ws_new.cell(row=row_index, column=col_index, value=value)
                
                # Apply header styles to the first row
                if row_index == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                else:
                    cell.alignment = center_alignment
                    cell.border = thin_border

        # Adjust column widths
        for col in ws_new.columns:
            max_length = 0
            column = col[0].column_letter  # Get column letter
            for cell in col:
                try:  # Check for cell value length
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws_new.column_dimensions[column].width = adjusted_width

        # Move the new sheet to the starting position and set as active
        wb._sheets.insert(0, wb._sheets.pop(-1))
        wb.active = ws_new
    except Exception as e:
        pass

    return wb

def downloadReconReport(request):
    
    req_data=json.loads(request.body)
    
    start_date,end_date=get_month_start_end_dates(req_data['formdata']['selected_month'])
        
    startdate=add530hrstoDateString(start_date).date()
    enddate=add530hrstoDateString(end_date).date()
    acc_type = req_data['formdata']['acc_type']

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment;'
    
    wb=load_workbook(filename = 'DSM_FinReconReport.xlsx')         
    ws=wb.active
    all_users=list(Registration.objects.filter(Q(end_date__isnull=True)|Q(end_date__gte=datetime.today())).order_by('fees_charges_name').values_list('fees_charges_name','fin_code'))

    disbursed_entities_obj=DisbursedEntities.objects.filter(pool_acctype = acc_type).all()
    week_start_enddates_obj = YearCalendar.objects.all()

    prev_month_year = getPrevYearMonth(req_data['formdata']['selected_month']) 
    
    closing_balances_qry =  ClosingBalances.objects.filter(Month_year = prev_month_year , Acc_type = acc_type )

    parent_folder = os.path.abspath(os.path.join(base_dir, os.pardir))
    directory = os.path.join(parent_folder, 'Files', 'ReconSummary' )
    #import pdb ; pdb.set_trace()
    
    if not os.path.exists(directory):
        # Create the directory if it doesn't exist
        os.makedirs(directory)
   
    summary_sheet_data = [['Fin Code','Entity Name' , 'Closing Balance']]

    
    for user in all_users:
        try:
            fincode=user[1].replace(" ", "")
            fincode = 'S0183'
            closing_balance_list = list(closing_balances_qry.filter(Fin_code = fincode).values_list('Closing_amount',flat=True))
            closing_balance = closing_balance_list[0] if len(closing_balance_list) else 0
            
            source=wb.active
            
            target=wb.copy_worksheet(source)
            entity_name = [character for character in user[0] if character.isalnum()]
            entity_name = "".join(entity_name).upper()
            # if two entities names are same then generate Name with some random number
            sheet_name=entity_name[:20] if entity_name[:20] not in wb.sheetnames else entity_name[:20]+str(random.randint(1,10))
        
            target.title=sheet_name
            ws=wb[sheet_name]
            
            ws['A1'].value='DETAILS OF '+acc_type+' PAYMENT AND DISBURSEMENT of '+ str(user[0]) +'  DURING ('+(startdate.strftime("%d.%m.%Y"))+'-'+enddate.strftime("%d.%m.%Y") +')'
            
            start_payable, start_receivable = 6 , 6
           
            
            all_paid_inrange_list1 , all_rcv_inrange_list = reco_for_user(fincode,start_date,end_date,acc_type)


            df_all_paid_inrange_list1 = pd.DataFrame(all_paid_inrange_list1)
            
            
            
            
            df_all_paid_inrange_list1.drop_duplicates(inplace=True)
            
            all_paid_inrange_list1 = df_all_paid_inrange_list1.values.tolist()

            i=0
            payable_out=0
            for x in range(start_payable,len(all_paid_inrange_list1)+start_payable):
                j=0
                for y in range(1,9):
                    ws.cell(row=x,column=y).value=all_paid_inrange_list1[i][j]
                    j+=1
                
                payable_out+=all_paid_inrange_list1[i][-1]
                i+=1 

            start_payable+=i
            
            # Receivables
            i=0
            receivable_out=0
            for x in range(start_receivable,len(all_rcv_inrange_list)+start_receivable):
                j=0
                for y in range(9,17):
                    ws.cell(row=x,column=y).value=all_rcv_inrange_list[i][j]
                    j+=1     
                receivable_out+=all_rcv_inrange_list[i][-1]
                i+=1 
            start_receivable+=i
            
            present_row='6'
            if start_payable == 6:
                ws['B7'].value='TOTAL'
                ws['D7'].value=0
                ws['G7'].value=0
                ws['H7'].value=0
            else:
                present_row=str(start_payable)
                sum_hvalue = sum(ws[f"H{row}"].value or 0 for row in range(6, start_payable))
                
                ws['B'+present_row].value='TOTAL'
                ws['D'+present_row].value="=SUM(D6:D"+str(start_payable-1)+")"
                ws['G'+present_row].value="=SUM(G6:G"+str(start_payable-1)+")"
                ws['H'+present_row].value= sum_hvalue
                
            if start_receivable == 6:
                receivable_row='6'
                ws['J7'].value='TOTAL'
                ws['M7'].value=0
                ws['N7'].value=0
                ws['P7'].value=0

            else:
                receivable_row=str(start_receivable)

                sum_mvalue = sum(ws[f"P{row}"].value or 0 for row in range(6, start_receivable))

                ws['J'+receivable_row].value='TOTAL'
                ws['M'+receivable_row].value="=SUM(M6:M"+str(start_receivable-1)+")"
                ws['N'+receivable_row].value="=SUM(N6:N"+str(start_receivable-1)+")"
                ws['P'+receivable_row].value = sum_mvalue
            
            # # Fetch cell values
            H_present = ws['H' + str(present_row)].value or 0  # Default to 0 if None
            M_receivable = ws['P' + str(receivable_row)].value or 0
            
            # # Compute net balance
            net_balance = H_present - M_receivable + closing_balance
            # net_balance = "=H"+present_row+"-M"+receivable_row+"+H2"
            if start_payable==6 and start_receivable==6:
                ws['F8'].value='Net Balance'
                ws['G8'].value=0
            else:
                ws['G'+str(start_payable+2)].value='Net Balance'
                ws['H'+str(start_payable+2)].value = net_balance
                
                ws['G'+str(start_payable+2)].fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type = "solid") 
            # this is to add last month closing balance
            ws['H2'] = closing_balance
            
            # summary sheet
            summary_sheet_data.append([fincode , user[0] , net_balance])
        except Exception as e:
            print(e)
            continue

    del wb['Sheet1']   #deleting temp sheet
    wb = prepareSummarySheet(wb , summary_sheet_data)

    in_filename=str(req_data['formdata']['acc_type'])+'_MonthReconciliation_'+str(req_data['formdata']['selected_month'])+'.xlsx'
    full_path=os.path.join(directory, in_filename)
    wb.save(full_path)
    wb.save(response)    
    wb.close()
    return response  
    
   



def downloadSummaryReconReport(request):
    try:
        req_data=json.loads(request.body)
        selected_month_year = req_data['formdata']['selected_month']
        start_date,end_date=get_month_start_end_dates(selected_month_year)
           
        enddate=add530hrstoDateString(end_date).date()

        # get summary 
        closing_balance_df = pd.DataFrame(ClosingBalances.objects.filter(Month_year = selected_month_year , Acc_type = req_data['formdata']['acc_type']).values('Fin_code' , 'Closing_amount') , columns = ['Fin_code' , 'Closing_amount'])

        registration_df = pd.DataFrame(Registration.objects.filter(Q(end_date__isnull=True)|Q(end_date__gte=datetime.today())).order_by('fees_charges_name').values_list('fees_charges_name','fin_code') , columns=['fees_charges_name','fin_code'])

        merged_df = pd.merge(closing_balance_df ,registration_df , left_on='Fin_code' , right_on='fin_code' , how='left')
        merged_df.drop(columns=['fin_code'] , inplace=True)
        merged_df.rename(columns= {'Fin_code' : 'Party Code' , 'fees_charges_name':'Name' ,'Closing_amount':'MO Balance'} , inplace=True)

        # merged_df['MO Balance']=merged_df['MO Balance'].apply(lambda x :format_indian_currency(x) )

        # add two extra columns like Fin Balance and Remarks
        merged_df['Fin Balance'] = None
        merged_df['Remarks'] = None
        # Specify the new order
        new_order = ["Party Code", "Name", "Fin Balance" ,"MO Balance" , "Remarks"]
        # Reorder the DataFrame
        merged_df = merged_df[new_order]
        parent_folder = os.path.abspath(os.path.join(base_dir, os.pardir))
        directory = os.path.join(parent_folder, 'Files', 'ReconSummary' )
        if not os.path.exists(directory):
                # Create the directory if it doesn't exist
                os.makedirs(directory)
        in_filename='recon_summary&'+str(selected_month_year)+'.csv'
        full_path=os.path.join(directory, in_filename)
        temp_out=open(full_path,'w')
        temp_out.write(' \n ')
        temp_out.write('Reconciliation Summary as on '+ enddate.strftime('%d.%m.%Y') +' \n')
        temp_out.close()
        merged_df.to_csv(full_path, mode='a',index=False )  
  
        return FileResponse(open(full_path,'rb'),content_type='text/csv') 
    
    except Exception as e:
        print(e)
        return HttpResponse(extractdb_errormsg(e),status=400)

def downloadReconUploadStatus(request):
    try:
        req_data=json.loads(request.body)['formdata']
        acc_type = req_data['acc_type']
        fin_year = req_data['fin_year']
        quarter = req_data['quarter']
        
        if not checkBillsNotified(acc_type , fin_year , quarter):
            return HttpResponse( 'Bills not notified , Please wait', status = 404)
        registration_df = pd.DataFrame(Registration.objects.filter(Q(end_date__isnull=True)|Q(end_date__gte=datetime.today())).order_by('fees_charges_name').values_list('fees_charges_name','fin_code') , columns=['fees_charges_name','fin_code'])

        signed_df = pd.DataFrame(ReconUploadStatus.objects.filter(Acc_type = acc_type ,Fin_year = fin_year , Quarter = quarter ).values('Fin_code','Upload_status','Uploaded_time') , columns=['Fin_code','Upload_status','Uploaded_time'] )

        merged_df = pd.merge(registration_df ,signed_df , left_on='fin_code' , right_on='Fin_code' , how='left')
        merged_df.drop(columns=['Fin_code' ] , inplace=True)
        merged_df.fillna('',inplace=True)
        merged_df.rename(columns= {'fin_code' : 'Party Code' , 'fees_charges_name':'Name' ,'Upload_status':'Upload Status' ,'Uploaded_time' :'Uploaded Time'} , inplace=True)
        # Specify the new order
        new_order = ["Party Code", "Name", "Upload Status" ,"Uploaded Time"]
        # Reorder the DataFrame
        merged_df = merged_df[new_order]
        parent_folder = os.path.abspath(os.path.join(base_dir, os.pardir))
        directory = os.path.join(parent_folder, 'Files', 'ReconSummary' )
        if not os.path.exists(directory):
            # Create the directory if it doesn't exist
            os.makedirs(directory)
        in_filename='recon_summary&'+fin_year+'&'+quarter+'.csv'
        full_path=os.path.join(directory, in_filename)
        temp_out=open(full_path,'w')
        temp_out.write(' \n ')
        temp_out.write('Reconciliation Summary for the fin year '+ req_data['fin_year'] +' and for the quarter '+ req_data['quarter'] +' \n')
        temp_out.close()
        merged_df.to_csv(full_path, mode='a',index=False )  
        return FileResponse(open(full_path,'rb'),content_type='text/csv') 
    
    except Exception as e:
        return HttpResponse(extractdb_errormsg(e),status=400)
